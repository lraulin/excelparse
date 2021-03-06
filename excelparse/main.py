import os
import pandas as pd
import requests
import urllib3
from shutil import copyfile
import sys
import requests
from openpyxl import Workbook, load_workbook
import json
import time
from sys import argv

TIMEOUT = 15
SHEET_NAME = "Source-Live Websites w Owners"
OUTPUT_FILE = "url_responses.json"


def create_backup(file_name):
    backup_filename = f"{file_name}.bak~{time.time()}"
    copyfile(file_name, backup_filename)

def check_for_redirects(url):
    print(f"Checking {url}...")
    result = {
        "url": url,
        "status_code": None,
        "message": None,
        "r_url": None
    }
    try:
        r = requests.get(url, allow_redirects=False, timeout=TIMEOUT)
        result["status_code"] = r.status_code
        print(f"Status Code: {r.status_code}")
        if 300 <= r.status_code < 400:
            re_res = requests.get(url, allow_redirects=True)
            result['r_url'] = re_res.url
            print(f"Location: {re_res.url}")
        else:
            result["message"] = '[no redirect]'
            print('No Redirect')
    except requests.exceptions.Timeout:
        print("Timeout")
        result["message"] = f'[timeout after {TIMEOUT} s]'
    except requests.exceptions.ConnectionError:
        print("Connection Error")
        result["message"] = '[connection error]'
    finally:
        return result


def write_list(my_list):
    with open('your_file.txt', 'w') as f:
        for item in my_list:
            f.write("%s\n" % item)

def check_websites(file_name, column_header):
    xl_file = pd.ExcelFile(file_name)
    data_frame = xl_file.parse('Source-Live Websites w Owners')
    column = data_frame.loc[:, column_header]
    urls = [x for x in column]
    statuses = [check_for_redirects(x) for x in urls]
    return statuses


def create_live_col(sheet, arr):
    sheet.insert_cols(5)
    sheet.cell(column=5, row=1, value='Live')
    for i in range(0, len(arr)):
        sheet.cell(column=5, row=i+2, value=arr[i])

def update_redirect_col(sheet, arr):
    for i in range(len(arr)):
        is_redirect = False if arr[i] == None else True
        url = "" if arr[i] == None else arr[i]
        sheet.cell(column=6, row=i+2, value=is_redirect)
        sheet.cell(column=7, row=i+2, value=url)


def modify_xlsx(data, file_name):
    live = []
    redirect = []
    for item in data:
        live.append(
            True if (item['status_code'] == 200 or item['r_url'] != None)
            else False)
        redirect.append(item['r_url'])

    # Save data
    wb = load_workbook(file_name)
    sheet = wb[SHEET_NAME]
    create_live_col(sheet, live)
    update_redirect_col(sheet, redirect)
    wb.save(file_name)


def save_json(data, file):
    with open(file, 'w') as f:
        json.dump(data, f)

def load_json(file):
    with open(file, 'r') as f:
        return json.load(f)


def main():
    file_name = argv[1]
    statuses = check_websites(file_name, 'Canonical URL')
    save_json(statuses, OUTPUT_FILE)
    create_backup(file_name)
    modify_xlsx(statuses, file_name)


if __name__ == "__main__":
    main()